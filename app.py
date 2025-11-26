"""
Aplicación web Flask para monitorear disponibilidad del Colosseo
"""

import sys
import io
import json
import os
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
from io import BytesIO

# Configurar encoding UTF-8 para Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

from api_client import ColosseoAPIClient, AvailabilityChecker
import storage_client

app = Flask(__name__)

# Configuración de tours disponibles
TOURS = {
    "24h-grupos": {
        "nombre": "24h Colosseo, Foro Romano y Palatino - GRUPOS",
        "guid": "a9a4b0f8-bf3c-4f22-afcd-196a27be04b9"
    },
    "arena": {
        "nombre": "Colosseo con ACCESO A LA ARENA",
        "guid": "8d1c991c-a15f-42bc-8cb5-bd738aa19c70"
    }
}


def consultar_tour_completo(client, tour_guid, meses_a_consultar):
    """
    Consulta un tour en múltiples meses

    Returns:
        tuple (dict con fechas y sus datos, lista de errores)
    """
    datos_totales = {}
    debug_info = []

    for month in meses_a_consultar:
        data, status, msg = client.fetch_calendar_data(
            guid=tour_guid,
            month=month
        )

        debug_info.append(f"{month}: status={status}, data={len(data) if data else 0}, msg={msg[:50] if msg else 'ok'}")

        if data:
            checker = AvailabilityChecker()
            normalized = checker.normalize_data(data)
            datos_totales.update(normalized)

    return datos_totales, debug_info


def obtener_timeslots_detallados(client, tour_guid, meses_a_consultar):
    """
    Obtiene todos los timeslots sin agregar, con horarios individuales

    Returns:
        list de timeslots con todos sus detalles
    """
    todos_timeslots = []

    for month in meses_a_consultar:
        data, status, msg = client.fetch_calendar_data(
            guid=tour_guid,
            month=month
        )

        if data and isinstance(data, list):
            for timeslot in data:
                if isinstance(timeslot, dict):
                    # Extraer información del timeslot
                    start_time = timeslot.get('startDateTime', '')
                    end_time = timeslot.get('endDateTime', '')
                    capacity = timeslot.get('capacity', 0)
                    original_capacity = timeslot.get('originalCapacity', 0)

                    if start_time:
                        # Parsear fecha y hora
                        try:
                            if 'T' in start_time:
                                fecha_hora = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                                # Convertir de UTC a hora italiana (CET/CEST = UTC+1)
                                fecha_hora = fecha_hora + timedelta(hours=1)
                                fecha = fecha_hora.strftime('%Y-%m-%d')
                                hora = fecha_hora.strftime('%H:%M')
                            else:
                                fecha = start_time
                                hora = "N/A"

                            ocupadas = original_capacity - capacity if original_capacity else 0
                            porcentaje_ocupado = (ocupadas / original_capacity * 100) if original_capacity > 0 else 0

                            todos_timeslots.append({
                                'fecha': fecha,
                                'hora': hora,
                                'start_datetime': start_time,
                                'end_datetime': end_time,
                                'capacidad': capacity,
                                'capacidad_original': original_capacity,
                                'ocupadas': ocupadas,
                                'porcentaje_ocupado': round(porcentaje_ocupado, 1),
                                'disponible': capacity > 0
                            })
                        except:
                            continue

    return todos_timeslots


def formatear_resultados_para_tabla(datos_totales):
    """
    Formatea los datos para mostrar en tabla HTML/Excel

    Returns:
        list de dicts con formato para tabla
    """
    resultados = []

    fechas_ordenadas = sorted(datos_totales.items())

    for fecha, info in fechas_ordenadas:
        capacidad = info.get('capacity', 0)
        capacidad_orig = info.get('originalCapacity', 0)
        ocupadas = capacidad_orig - capacidad if capacidad_orig else 0
        porcentaje_ocupado = (ocupadas / capacidad_orig * 100) if capacidad_orig > 0 else 0

        # Determinar estado
        if capacidad == 0:
            estado = "AGOTADO"
            nivel = "agotado"
        elif porcentaje_ocupado < 30:
            estado = "MUCHA DISPONIBILIDAD"
            nivel = "alta"
        elif porcentaje_ocupado < 70:
            estado = "DISPONIBILIDAD MODERADA"
            nivel = "media"
        else:
            estado = "POCA DISPONIBILIDAD"
            nivel = "baja"

        # Formatear fecha con día de semana
        try:
            fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")
            dia_semana = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"][fecha_obj.weekday()]
            fecha_formateada = f"{fecha} ({dia_semana})"
        except:
            fecha_formateada = fecha
            dia_semana = ""

        resultados.append({
            "fecha": fecha,
            "fecha_formateada": fecha_formateada,
            "dia_semana": dia_semana,
            "plazas_disponibles": capacidad,
            "plazas_totales": capacidad_orig,
            "plazas_ocupadas": ocupadas,
            "porcentaje_ocupado": round(porcentaje_ocupado, 1),
            "estado": estado,
            "nivel": nivel
        })

    return resultados


def calcular_estadisticas_horarios(timeslots):
    """
    Calcula estadísticas sobre los horarios más demandados

    Returns:
        dict con estadísticas
    """
    from collections import defaultdict

    # Agrupar por hora
    por_hora = defaultdict(lambda: {'total': 0, 'agotados': 0, 'capacidad_total': 0, 'ocupadas_total': 0})
    por_fecha = defaultdict(list)

    for ts in timeslots:
        hora = ts['hora']
        fecha = ts['fecha']

        por_hora[hora]['total'] += 1
        por_hora[hora]['capacidad_total'] += ts['capacidad_original']
        por_hora[hora]['ocupadas_total'] += ts['ocupadas']

        if ts['capacidad'] == 0:
            por_hora[hora]['agotados'] += 1

        por_fecha[fecha].append(ts)

    # Calcular promedios y porcentajes
    estadisticas_hora = []
    for hora, stats in sorted(por_hora.items()):
        if stats['total'] > 0:
            porcentaje_agotado = (stats['agotados'] / stats['total']) * 100
            ocupacion_promedio = (stats['ocupadas_total'] / stats['capacidad_total'] * 100) if stats['capacidad_total'] > 0 else 0

            estadisticas_hora.append({
                'hora': hora,
                'total_timeslots': stats['total'],
                'timeslots_agotados': stats['agotados'],
                'porcentaje_agotado': round(porcentaje_agotado, 1),
                'ocupacion_promedio': round(ocupacion_promedio, 1)
            })

    # Ordenar por porcentaje de agotado (más demandados primero)
    estadisticas_hora.sort(key=lambda x: x['porcentaje_agotado'], reverse=True)

    # Calcular días hasta agotamiento
    hoy = datetime.now().date()
    dias_hasta_agotamiento = []

    for fecha, timeslots_fecha in por_fecha.items():
        try:
            fecha_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
            dias_diferencia = (fecha_obj - hoy).days

            agotados = sum(1 for ts in timeslots_fecha if ts['capacidad'] == 0)
            total = len(timeslots_fecha)

            if agotados > 0:
                dias_hasta_agotamiento.append({
                    'fecha': fecha,
                    'dias_adelantados': dias_diferencia,
                    'timeslots_agotados': agotados,
                    'total_timeslots': total,
                    'porcentaje_agotado': round((agotados / total) * 100, 1)
                })
        except:
            continue

    # Ordenar por días adelantados (más lejanos primero)
    dias_hasta_agotamiento.sort(key=lambda x: x['dias_adelantados'], reverse=True)

    return {
        'por_hora': estadisticas_hora[:10],  # Top 10 horas más demandadas
        'dias_agotamiento': dias_hasta_agotamiento[:20]  # Top 20 fechas
    }


@app.route('/')
def index():
    """Página principal"""
    return render_template('index.html', tours=TOURS)


@app.route('/api/consultar', methods=['POST'])
def consultar_disponibilidad():
    """
    Endpoint para consultar disponibilidad

    Recibe:
        - cookies: JSON con cookies
        - tours: Lista de IDs de tours a consultar
        - meses: Número de meses a consultar (default: 6)

    Returns:
        JSON con resultados
    """
    try:
        data = request.json

        # Validar cookies
        cookies_json = data.get('cookies')
        if not cookies_json:
            return jsonify({"error": "No se proporcionaron cookies"}), 400

        # Parsear cookies
        try:
            cookies = json.loads(cookies_json)
        except json.JSONDecodeError:
            return jsonify({"error": "Formato de cookies inválido. Debe ser JSON válido"}), 400

        # Obtener tours seleccionados
        tours_seleccionados = data.get('tours', list(TOURS.keys()))
        num_meses = data.get('meses', 6)

        # Calcular meses a consultar
        hoy = datetime.now()
        meses_a_consultar = []
        for i in range(num_meses):
            fecha = hoy + timedelta(days=30*i)
            mes_str = f"{fecha.year}-{fecha.month:02d}"
            if mes_str not in meses_a_consultar:
                meses_a_consultar.append(mes_str)

        # Crear cliente con cookies
        client = ColosseoAPIClient()
        client.cookies = cookies
        client.create_session_from_cookies(cookies)

        # Consultar cada tour
        resultados = {}
        errores = []

        for tour_key in tours_seleccionados:
            if tour_key not in TOURS:
                continue

            tour_info = TOURS[tour_key]

            try:
                # Consultar disponibilidad agregada por fecha
                datos, debug_info = consultar_tour_completo(
                    client,
                    tour_info['guid'],
                    meses_a_consultar
                )

                # Agregar debug info a errores para ver qué pasa
                if not datos:
                    errores.append(f"{tour_key} sin datos: {'; '.join(debug_info[:3])}")

                # Obtener timeslots detallados
                timeslots = obtener_timeslots_detallados(
                    client,
                    tour_info['guid'],
                    meses_a_consultar
                )
            except Exception as e:
                errores.append(f"{tour_key}: {str(e)}")
                continue

            if datos:
                # Formatear resultados agregados
                fechas_formateadas = formatear_resultados_para_tabla(datos)
                total_plazas = sum(f['plazas_disponibles'] for f in fechas_formateadas)

                # Agrupar timeslots por fecha
                timeslots_por_fecha = {}
                for ts in timeslots:
                    fecha = ts['fecha']
                    if fecha not in timeslots_por_fecha:
                        timeslots_por_fecha[fecha] = []
                    timeslots_por_fecha[fecha].append(ts)

                # Calcular estadísticas
                estadisticas = calcular_estadisticas_horarios(timeslots)

                resultados[tour_key] = {
                    "nombre": tour_info['nombre'],
                    "guid": tour_info['guid'],
                    "total_fechas": len(fechas_formateadas),
                    "total_plazas": total_plazas,
                    "fechas": fechas_formateadas,
                    "timeslots_por_fecha": timeslots_por_fecha,
                    "estadisticas": estadisticas
                }

        if not resultados:
            error_msg = "No se pudieron obtener datos. Verifica las cookies"
            if errores:
                error_msg += f". Errores: {'; '.join(errores)}"
            # Debug info
            error_msg += f". Tours intentados: {tours_seleccionados}. Meses: {meses_a_consultar}. Cookies recibidas: {len(cookies)}"
            return jsonify({"error": error_msg}), 400

        return jsonify({
            "success": True,
            "meses_consultados": meses_a_consultar,
            "resultados": resultados,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

    except Exception as e:
        return jsonify({"error": f"Error al procesar: {str(e)}"}), 500


@app.route('/api/exportar-excel', methods=['POST'])
def exportar_excel():
    """
    Exporta los resultados a Excel

    Recibe:
        - resultados: Datos de la consulta

    Returns:
        Archivo Excel
    """
    try:
        data = request.json
        resultados = data.get('resultados', {})

        if not resultados:
            return jsonify({"error": "No hay datos para exportar"}), 400

        # Crear archivo Excel en memoria
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja de resumen
            resumen_data = []
            for tour_key, tour_data in resultados.items():
                resumen_data.append({
                    "Tour": tour_data['nombre'],
                    "Fechas Disponibles": tour_data['total_fechas'],
                    "Plazas Totales": tour_data['total_plazas']
                })

            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)

            # Una hoja por tour
            for tour_key, tour_data in resultados.items():
                fechas = tour_data['fechas']

                df_data = []
                for fecha_info in fechas:
                    df_data.append({
                        "Fecha": fecha_info['fecha'],
                        "Día": fecha_info['dia_semana'],
                        "Plazas Disponibles": fecha_info['plazas_disponibles'],
                        "Plazas Totales": fecha_info['plazas_totales'],
                        "% Ocupado": fecha_info['porcentaje_ocupado'],
                        "Estado": fecha_info['estado']
                    })

                df = pd.DataFrame(df_data)

                # Nombre de hoja (max 31 caracteres)
                sheet_name = tour_data['nombre'][:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        output.seek(0)

        # Generar nombre de archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"colosseo_disponibilidad_{timestamp}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({"error": f"Error al exportar: {str(e)}"}), 500


@app.route('/api/cargar-cookies-archivo', methods=['POST'])
def cargar_cookies_archivo():
    """
    Carga cookies desde cookies_colosseo.json si existe

    Returns:
        JSON con cookies
    """
    try:
        client = ColosseoAPIClient()
        if client.load_cookies():
            return jsonify({
                "success": True,
                "cookies": json.dumps(client.cookies, indent=2)
            })
        else:
            return jsonify({
                "success": False,
                "error": "No se encontró el archivo cookies_colosseo.json"
            }), 404

    except Exception as e:
        return jsonify({"error": f"Error al cargar cookies: {str(e)}"}), 500


@app.route('/api/guardar-cookies', methods=['POST'])
def guardar_cookies():
    """
    Guarda cookies en cookies_colosseo.json

    Returns:
        JSON con resultado
    """
    try:
        data = request.json
        cookies_json = data.get('cookies')

        if not cookies_json:
            return jsonify({"error": "No se proporcionaron cookies"}), 400

        # Parsear cookies
        try:
            cookies = json.loads(cookies_json)
        except json.JSONDecodeError:
            return jsonify({"error": "Formato de cookies inválido"}), 400

        # Guardar en archivo
        client = ColosseoAPIClient()
        client.cookies = cookies
        client.save_cookies(cookies)

        return jsonify({
            "success": True,
            "message": f"Cookies guardadas correctamente ({len(cookies)} cookies)"
        })

    except Exception as e:
        return jsonify({"error": f"Error al guardar: {str(e)}"}), 500


@app.route('/api/guardar-historico', methods=['POST'])
def guardar_historico():
    """
    Guarda datos históricos de disponibilidad en Excel formato matriz para análisis de tendencias
    Formato: Cada fila = fecha+hora, cada columna = timestamp de consulta

    Returns:
        JSON con resultado
    """
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import os

        data = request.json
        resultados = data.get('resultados', {})

        if not resultados:
            return jsonify({"error": "No hay datos para guardar"}), 400

        # Archivo de histórico
        filename = 'historico_disponibilidad.xlsx'
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

        # Crear o cargar workbook
        if os.path.exists(filename):
            wb = load_workbook(filename)
        else:
            wb = Workbook()
            # Eliminar hoja por defecto
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

        # Procesar cada tour
        for tour_key, tour_data in resultados.items():
            tour_nombre = tour_data['nombre']
            timeslots_por_fecha = tour_data.get('timeslots_por_fecha', {})

            # Nombre de hoja (máx 31 caracteres)
            sheet_name = f"{tour_key[:25]}"

            # Crear estructura de datos: {(fecha, hora): {timestamp: disponibilidad}}
            matriz_datos = {}
            for fecha, timeslots in sorted(timeslots_por_fecha.items()):
                for ts in sorted(timeslots, key=lambda x: x.get('hora', '')):
                    hora = ts.get('hora', 'N/A')
                    key = (fecha, hora)

                    if key not in matriz_datos:
                        matriz_datos[key] = {
                            'capacidad_total': ts.get('capacidad_original', 0)
                        }

                    matriz_datos[key][timestamp] = ts.get('capacidad', 0)

            # === CREAR/ACTUALIZAR HOJA ===
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)

                # Headers fijos
                ws['A1'] = 'Fecha'
                ws['B1'] = 'Hora'
                ws['C1'] = 'Capacidad Total'

                # Estilo headers
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)

                for col in ['A1', 'B1', 'C1']:
                    ws[col].fill = header_fill
                    ws[col].font = header_font
                    ws[col].alignment = Alignment(horizontal='center', vertical='center')

                # Escribir filas (fecha + hora)
                row = 2
                for (fecha, hora), datos in sorted(matriz_datos.items()):
                    ws[f'A{row}'] = fecha
                    ws[f'B{row}'] = hora
                    ws[f'C{row}'] = datos['capacidad_total']
                    row += 1

                # Ajustar anchos
                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 8
                ws.column_dimensions['C'].width = 14

                # Primera columna de datos (timestamp actual)
                col_num = 4
            else:
                ws = wb[sheet_name]

                # Leer filas existentes y crear mapeo (fecha, hora) -> número de fila
                filas_existentes = {}
                for row in range(2, ws.max_row + 1):
                    fecha_cell = ws[f'A{row}'].value
                    hora_cell = ws[f'B{row}'].value
                    if fecha_cell and hora_cell:
                        key = (str(fecha_cell), str(hora_cell))
                        filas_existentes[key] = row

                # Identificar horarios nuevos que no existen en el archivo
                horarios_nuevos = []
                for (fecha, hora) in sorted(matriz_datos.keys()):
                    key = (fecha, hora)
                    if key not in filas_existentes:
                        horarios_nuevos.append(key)

                # Agregar horarios nuevos al final
                if horarios_nuevos:
                    next_row = ws.max_row + 1
                    for (fecha, hora) in horarios_nuevos:
                        ws[f'A{next_row}'] = fecha
                        ws[f'B{next_row}'] = hora
                        ws[f'C{next_row}'] = matriz_datos[(fecha, hora)]['capacidad_total']
                        filas_existentes[(fecha, hora)] = next_row
                        next_row += 1

                # Encontrar la siguiente columna disponible
                col_num = ws.max_column + 1

            # Agregar nueva columna con timestamp
            col_letter = get_column_letter(col_num)

            # Header de la nueva columna (timestamp)
            ws[f'{col_letter}1'] = timestamp
            ws[f'{col_letter}1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            ws[f'{col_letter}1'].font = Font(color='FFFFFF', bold=True)
            ws[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[col_letter].width = 12

            # Llenar datos de disponibilidad en las filas correspondientes
            if sheet_name not in wb.sheetnames or col_num == 4:
                # Primera vez: iterar en orden
                row = 2
                for (fecha, hora), datos in sorted(matriz_datos.items()):
                    disponibilidad = datos.get(timestamp, 0)
                    cell = ws[f'{col_letter}{row}']
                    cell.value = disponibilidad
                    cell.alignment = Alignment(horizontal='center')

                    # Formato condicional por color
                    capacidad_total = datos['capacidad_total']
                    if capacidad_total > 0:
                        porcentaje = (disponibilidad / capacidad_total) * 100
                        if disponibilidad == 0:
                            cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                            cell.font = Font(color='FFFFFF', bold=True)
                        elif porcentaje < 30:
                            cell.fill = PatternFill(start_color='FFE066', end_color='FFE066', fill_type='solid')
                        elif porcentaje > 70:
                            cell.fill = PatternFill(start_color='95E1D3', end_color='95E1D3', fill_type='solid')

                    row += 1
            else:
                # Actualizaciones posteriores: usar mapeo de filas existentes
                for (fecha, hora), datos in matriz_datos.items():
                    key = (fecha, hora)
                    if key in filas_existentes:
                        row_num = filas_existentes[key]
                        disponibilidad = datos.get(timestamp, 0)
                        cell = ws[f'{col_letter}{row_num}']
                        cell.value = disponibilidad
                        cell.alignment = Alignment(horizontal='center')

                        # Formato condicional por color
                        capacidad_total = datos['capacidad_total']
                        if capacidad_total > 0:
                            porcentaje = (disponibilidad / capacidad_total) * 100
                            if disponibilidad == 0:
                                cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                                cell.font = Font(color='FFFFFF', bold=True)
                            elif porcentaje < 30:
                                cell.fill = PatternFill(start_color='FFE066', end_color='FFE066', fill_type='solid')
                            elif porcentaje > 70:
                                cell.fill = PatternFill(start_color='95E1D3', end_color='95E1D3', fill_type='solid')

                # Para filas que ya no tienen datos (horarios pasados), dejar vacías con marca
                for row in range(2, ws.max_row + 1):
                    fecha_cell = ws[f'A{row}'].value
                    hora_cell = ws[f'B{row}'].value
                    if fecha_cell and hora_cell:
                        key = (str(fecha_cell), str(hora_cell))
                        if key not in matriz_datos:
                            # Este horario ya no está en la consulta actual (pasó o se eliminó)
                            cell = ws[f'{col_letter}{row}']
                            cell.value = '-'  # Marca para indicar que no está disponible
                            cell.alignment = Alignment(horizontal='center')
                            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                            cell.font = Font(color='999999', italic=True)

        # Contar timeslots
        total_timeslots = sum(
            len(tour_data.get('timeslots_por_fecha', {}).get(fecha, []))
            for tour_data in resultados.values()
            for fecha in tour_data.get('timeslots_por_fecha', {})
        )

        # Guardar archivo - local o Supabase
        if storage_client.is_configured():
            # Guardar en Supabase Storage
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            upload_result = storage_client.upload_file(
                output.getvalue(),
                filename,
                folder='historico'
            )

            if upload_result['success']:
                return jsonify({
                    "success": True,
                    "message": f"Historico guardado en la nube ({total_timeslots} horarios)",
                    "filename": filename,
                    "timestamp": timestamp,
                    "url": upload_result.get('url', '')
                })
            else:
                # Fallback a local si falla Supabase
                wb.save(filename)
                return jsonify({
                    "success": True,
                    "message": f"Historico guardado localmente ({total_timeslots} horarios)",
                    "filename": filename,
                    "timestamp": timestamp,
                    "warning": "No se pudo guardar en la nube"
                })
        else:
            # Guardar localmente
            wb.save(filename)
            return jsonify({
                "success": True,
                "message": f"Historico actualizado ({total_timeslots} horarios)",
                "filename": filename,
                "timestamp": timestamp
            })

    except Exception as e:
        return jsonify({"error": f"Error al guardar historico: {str(e)}"}), 500


@app.route('/api/descargar-historico', methods=['GET'])
def descargar_historico():
    """
    Descarga el archivo histórico desde Supabase o local.

    Returns:
        Archivo Excel o JSON con URL
    """
    try:
        if storage_client.is_configured():
            # Obtener URL desde Supabase
            result = storage_client.get_historico_url()
            if result['success']:
                return jsonify({
                    "success": True,
                    "url": result['url']
                })
            else:
                return jsonify({"error": "Archivo no encontrado en la nube"}), 404
        else:
            # Descargar archivo local
            filename = 'historico_disponibilidad.xlsx'
            if os.path.exists(filename):
                return send_file(
                    filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
            else:
                return jsonify({"error": "Archivo no encontrado"}), 404

    except Exception as e:
        return jsonify({"error": f"Error: {str(e)}"}), 500


@app.route('/api/storage-status', methods=['GET'])
def storage_status():
    """Verifica el estado del almacenamiento"""
    return jsonify({
        "supabase_configured": storage_client.is_configured(),
        "mode": "cloud" if storage_client.is_configured() else "local"
    })


# Para Vercel - exportar la app
application = app

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'

    print("\n" + "=" * 70)
    print("COLOSSEO MONITOR - APLICACION WEB")
    print("=" * 70)
    print(f"\nIniciando servidor en puerto {port}...")
    print("=" * 70 + "\n")

    app.run(debug=debug, host='0.0.0.0', port=port)
