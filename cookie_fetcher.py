"""
Cookie Fetcher para Railway usando undetected-chromedriver.
Obtiene cookies del Colosseo pasando Octofence y las sube a Supabase.
"""

import os
import sys
import json
import time
import subprocess
from datetime import datetime

# Iniciar Xvfb para display virtual
def start_xvfb():
    try:
        subprocess.Popen(['Xvfb', ':99', '-screen', '0', '1920x1080x24'],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(2)
        os.environ['DISPLAY'] = ':99'
        print("[Xvfb] Display virtual iniciado")
    except Exception as e:
        print(f"[Xvfb] Error: {e}")

# Configuracion
SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', '')
TOUR_URL = "https://ticketing.colosseo.it/en/eventi/24h-colosseo-foro-romano-palatino-gruppi/"

# Webshare API Key para obtener proxies automáticamente
WEBSHARE_API_KEY = os.environ.get('WEBSHARE_API_KEY', '')

# Fallback: Proxy manual si no hay API key
PROXY_HOST = os.environ.get('PROXY_HOST', '')
PROXY_PORT = os.environ.get('PROXY_PORT', '')
PROXY_USER = os.environ.get('PROXY_USER', '')
PROXY_PASS = os.environ.get('PROXY_PASS', '')

def fetch_webshare_proxies():
    """Obtiene lista de proxies desde la API de Webshare"""
    import requests

    if not WEBSHARE_API_KEY:
        print("[Webshare] No hay API key configurada")
        return []

    try:
        print("[Webshare] Obteniendo lista de proxies...")
        response = requests.get(
            "https://proxy.webshare.io/api/v2/proxy/list/?mode=direct&page=1&page_size=100",
            headers={"Authorization": f"Token {WEBSHARE_API_KEY}"},
            timeout=30
        )

        if response.status_code != 200:
            print(f"[Webshare] Error API: {response.status_code} - {response.text[:200]}")
            return []

        data = response.json()
        proxies = []

        for proxy in data.get('results', []):
            if proxy.get('valid', False):
                proxies.append({
                    'host': proxy['proxy_address'],
                    'port': str(proxy['port']),
                    'user': proxy['username'],
                    'pass': proxy['password'],
                    'country': proxy.get('country_code', 'XX')
                })

        print(f"[Webshare] {len(proxies)} proxies válidos obtenidos")
        return proxies

    except Exception as e:
        print(f"[Webshare] Error obteniendo proxies: {e}")
        return []

def parse_proxy_list():
    """Obtiene lista de proxies: primero de Webshare API, luego fallback manual"""
    proxies = []

    # Intentar obtener de Webshare API
    if WEBSHARE_API_KEY:
        proxies = fetch_webshare_proxies()

    # Fallback: usar proxy manual si no hay proxies de API
    if not proxies and PROXY_HOST and PROXY_PORT:
        print("[Proxy] Usando proxy manual como fallback")
        proxies.append({
            'host': PROXY_HOST,
            'port': PROXY_PORT,
            'user': PROXY_USER,
            'pass': PROXY_PASS,
            'country': 'manual'
        })

    return proxies

# Lista global de proxies (se carga al inicio)
PROXIES = []
current_proxy_index = 0

def initialize_proxies():
    """Inicializa la lista de proxies (llamar al inicio de main)"""
    global PROXIES
    PROXIES = parse_proxy_list()
    return PROXIES


def create_proxy_extension(proxy_host, proxy_port, proxy_user, proxy_pass):
    """Crea una extension de Chrome para autenticacion de proxy"""
    import zipfile
    import tempfile

    manifest_json = """
    {
        "version": "1.0.0",
        "manifest_version": 2,
        "name": "Chrome Proxy",
        "permissions": [
            "proxy",
            "tabs",
            "unlimitedStorage",
            "storage",
            "<all_urls>",
            "webRequest",
            "webRequestBlocking"
        ],
        "background": {
            "scripts": ["background.js"]
        },
        "minimum_chrome_version":"22.0.0"
    }
    """

    background_js = """
    var config = {
        mode: "fixed_servers",
        rules: {
            singleProxy: {
                scheme: "http",
                host: "%s",
                port: parseInt(%s)
            },
            bypassList: ["localhost"]
        }
    };

    chrome.proxy.settings.set({value: config, scope: "regular"}, function() {});

    function callbackFn(details) {
        return {
            authCredentials: {
                username: "%s",
                password: "%s"
            }
        };
    }

    chrome.webRequest.onAuthRequired.addListener(
        callbackFn,
        {urls: ["<all_urls>"]},
        ['blocking']
    );
    """ % (proxy_host, proxy_port, proxy_user, proxy_pass)

    # Crear extension temporal
    ext_dir = tempfile.mkdtemp()
    ext_path = os.path.join(ext_dir, 'proxy_auth.zip')

    with zipfile.ZipFile(ext_path, 'w') as zp:
        zp.writestr("manifest.json", manifest_json)
        zp.writestr("background.js", background_js)

    return ext_path


def get_current_proxy():
    """Obtiene el proxy actual de la lista"""
    global current_proxy_index
    if not PROXIES:
        return None
    return PROXIES[current_proxy_index % len(PROXIES)]


def rotate_proxy():
    """Rota al siguiente proxy de la lista"""
    global current_proxy_index
    if not PROXIES:
        return None

    current_proxy_index = (current_proxy_index + 1) % len(PROXIES)
    proxy = PROXIES[current_proxy_index]
    country = proxy.get('country', 'XX')
    print(f"[Proxy] Rotando a proxy {current_proxy_index + 1}/{len(PROXIES)}: {proxy['host']}:{proxy['port']} ({country})")
    return proxy


def setup_driver(proxy_override=None):
    """Configura undetected-chromedriver con network logging y proxy"""
    import undetected_chromedriver as uc

    options = uc.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-blink-features=AutomationControlled')

    # Usar proxy especificado o el actual de la lista
    proxy = proxy_override or get_current_proxy()

    if proxy:
        print(f"[Proxy] Configurando proxy residencial: {proxy['host']}:{proxy['port']}")

        if proxy.get('user') and proxy.get('pass'):
            # Crear extension para autenticacion
            print(f"[Proxy] Creando extension de autenticacion...")
            ext_path = create_proxy_extension(proxy['host'], proxy['port'], proxy['user'], proxy['pass'])
            options.add_extension(ext_path)
            print(f"[Proxy] Extension cargada")
        else:
            # Sin autenticacion
            options.add_argument(f"--proxy-server=http://{proxy['host']}:{proxy['port']}")
    else:
        print("[Proxy] Sin proxy configurado - usando IP directa (puede ser bloqueado)")

    # Habilitar logging de red para capturar cookies HttpOnly
    options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})

    # NO usar headless - usar Xvfb en su lugar
    driver = uc.Chrome(options=options, use_subprocess=True)

    print("[Driver] undetected-chromedriver iniciado con network logging")
    return driver


def wait_for_octofence(driver, timeout=120):
    """Espera a que Octofence termine de verificar"""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    print(f"[Wait] Esperando Octofence (max {timeout}s)...")
    start = time.time()

    while time.time() - start < timeout:
        try:
            # Verificar si pasamos Octofence
            title = driver.title.lower()
            url = driver.current_url

            # Si el titulo tiene el nombre del tour, pasamos
            if 'colosseo' in title and 'waiting' not in title:
                # Verificar elementos de la pagina
                try:
                    driver.find_element(By.CSS_SELECTOR, '.ui-datepicker, .tariff-option, input[name="slot"]')
                    print(f"[Wait] Octofence pasado en {int(time.time()-start)}s")
                    return True
                except:
                    pass

            # Mostrar progreso
            elapsed = int(time.time() - start)
            if elapsed % 10 == 0:
                print(f"[Wait] Verificando... ({elapsed}s) - Title: {title[:50]}")

            time.sleep(2)

        except Exception as e:
            print(f"[Wait] Error: {e}")
            time.sleep(2)

    print("[Wait] Timeout esperando Octofence")
    return False


def accept_cookies_banner(driver):
    """Acepta el banner de cookies si aparece"""
    from selenium.webdriver.common.by import By

    print("[Banner] Buscando banner de cookies...")
    try:
        # Buscar boton de aceptar cookies
        accept_selectors = [
            "button#cookie_action_close_header",
            "a.cli_action_button.cli-accept-all-btn",
            "button.cli-accept-btn",
            "[data-cli_action='accept_all']",
            ".cookie-law-info-accept"
        ]

        for selector in accept_selectors:
            try:
                btn = driver.find_element(By.CSS_SELECTOR, selector)
                if btn.is_displayed():
                    driver.execute_script("arguments[0].click();", btn)
                    print(f"[Banner] Cookies aceptadas con: {selector}")
                    time.sleep(1)
                    return True
            except:
                continue

        # Intentar con JavaScript directo
        driver.execute_script("""
            var btns = document.querySelectorAll('button, a');
            for (var b of btns) {
                var txt = b.textContent.toLowerCase();
                if (txt.includes('accept') || txt.includes('aceptar') || txt.includes('accetta')) {
                    b.click();
                    return;
                }
            }
        """)
        print("[Banner] Intentado aceptar via JS")
        time.sleep(1)
        return True

    except Exception as e:
        print(f"[Banner] No encontrado o error: {e}")
        return False


def complete_booking_flow(driver):
    """Completa el flujo de reserva para generar cookies de sesion"""
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    print("[Flow] Iniciando flujo de reserva...")

    try:
        # Primero aceptar banner de cookies
        accept_cookies_banner(driver)
        time.sleep(2)

        # ESTRATEGIA 1: Ir directamente al carrito para forzar sesion
        print("[Flow] Estrategia 1: Visitar carrito para crear sesion...")
        try:
            driver.get("https://ticketing.colosseo.it/en/cart/")
            time.sleep(5)
            print(f"[Flow] URL carrito: {driver.current_url}")
        except Exception as e:
            print(f"[Flow] Error visitando carrito: {e}")

        # Volver a la pagina del tour
        print("[Flow] Volviendo a pagina del tour...")
        driver.get("https://ticketing.colosseo.it/en/eventi/24h-colosseo-foro-romano-palatino-gruppi/")
        time.sleep(5)

        # Esperar a que el calendario este visible
        print("[Flow] Esperando calendario...")
        wait = WebDriverWait(driver, 30)

        # 1. Click en dia disponible del calendario usando JavaScript
        print("[Flow] Buscando dia en calendario...")
        try:
            # Esperar a que el calendario se cargue completamente
            time.sleep(3)

            # Scroll al calendario primero
            driver.execute_script("window.scrollTo(0, 300);")
            time.sleep(2)

            # Intentar encontrar un dia disponible con multiples selectores
            day_selectors = [
                ".ui-datepicker-calendar td:not(.ui-datepicker-unselectable) a",
                ".ui-datepicker-calendar td.available a",
                ".ui-datepicker-calendar td a.ui-state-default",
                "td.abc-availability-available a"
            ]

            day = None
            for selector in day_selectors:
                try:
                    day = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
                    if day:
                        print(f"[Flow] Dia encontrado con selector: {selector}")
                        break
                except:
                    continue

            if day:
                # Scroll al elemento y click con JS
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", day)
                time.sleep(1)
                driver.execute_script("arguments[0].click();", day)
                print("[Flow] Dia seleccionado")
                time.sleep(4)
            else:
                # Intentar con JavaScript directo
                print("[Flow] Intentando seleccionar dia via JS...")
                result = driver.execute_script("""
                    var days = document.querySelectorAll('.ui-datepicker-calendar td:not(.ui-datepicker-unselectable) a');
                    if (days.length > 0) {
                        days[0].click();
                        return 'Clicked day ' + days[0].textContent;
                    }
                    return 'No days found';
                """)
                print(f"[Flow] JS result: {result}")
                time.sleep(4)
        except Exception as e:
            print(f"[Flow] No se pudo seleccionar dia: {e}")

        # 2. Esperar a que carguen los horarios
        print("[Flow] Esperando horarios...")
        time.sleep(4)

        # 3. Click en horario usando JavaScript
        print("[Flow] Buscando horario...")
        try:
            # Buscar el primer slot disponible
            slots = driver.find_elements(By.CSS_SELECTOR, "input[name='slot']")
            print(f"[Flow] Encontrados {len(slots)} slots")
            if slots:
                slot = slots[0]
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", slot)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", slot)
                print("[Flow] Horario seleccionado")
                time.sleep(3)
        except Exception as e:
            print(f"[Flow] No se pudo seleccionar horario: {e}")

        # 4. Seleccionar tarifa si existe
        print("[Flow] Buscando tarifas...")
        try:
            tariffs = driver.find_elements(By.CSS_SELECTOR, "input[name='tariff'], .tariff-option input[type='radio']")
            print(f"[Flow] Encontradas {len(tariffs)} tarifas")
            if tariffs:
                tariff = tariffs[0]
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tariff)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", tariff)
                print("[Flow] Tarifa seleccionada")
                time.sleep(2)
        except Exception as e:
            print(f"[Flow] No se encontraron tarifas: {e}")

        # 5. Incrementar cantidad usando JavaScript - SOLO 1 entrada para evitar problemas
        print("[Flow] Incrementando cantidad (1 entrada)...")
        try:
            # Buscar el primer boton + que NO sea de un item ya agregado
            increment_result = driver.execute_script("""
                var plusBtns = document.querySelectorAll('button[data-dir="up"]');
                if (plusBtns.length > 0) {
                    // Click en el primer boton +
                    plusBtns[0].scrollIntoView({block: 'center'});
                    plusBtns[0].click();
                    return 'Clicked first + button';
                }
                return 'No + buttons found';
            """)
            print(f"[Flow] Increment result: {increment_result}")
            time.sleep(3)  # Esperar a que se actualice la UI

            # Verificar que se incremento
            qty_check = driver.execute_script("""
                var inputs = document.querySelectorAll('input[type="number"]');
                var total = 0;
                for (var inp of inputs) {
                    total += parseInt(inp.value) || 0;
                }
                return total;
            """)
            print(f"[Flow] Cantidad total seleccionada: {qty_check}")

        except Exception as e:
            print(f"[Flow] No se pudo incrementar: {e}")

        # 5b. CRITICO: Hacer click en boton "Add to Cart" / "Book"
        print("[Flow] Buscando boton ADD TO CART...")
        try:
            # Primero mostrar el estado actual de la pagina
            page_state = driver.execute_script("""
                var result = {};

                // Verificar si hay items seleccionados
                var qtyInputs = document.querySelectorAll('input[type="number"]');
                var totalQty = 0;
                for (var inp of qtyInputs) {
                    totalQty += parseInt(inp.value) || 0;
                }
                result.totalQuantity = totalQty;

                // Verificar formularios de reserva
                var bookingForms = document.querySelectorAll('form[action*="cart"], form[action*="book"], form.booking-form, form[id*="book"]');
                result.bookingForms = bookingForms.length;

                // Buscar botones de submit dentro de formularios
                var submitBtns = document.querySelectorAll('form input[type="submit"], form button[type="submit"]');
                var submitTexts = [];
                for (var btn of submitBtns) {
                    submitTexts.push((btn.value || btn.textContent || '').substring(0, 30));
                }
                result.submitButtons = submitTexts;

                return JSON.stringify(result);
            """)
            print(f"[Flow] Page state: {page_state}")

            add_to_cart_result = driver.execute_script("""
                // PASO 0: PRIORIDAD MAXIMA - Buscar boton "Confirm" (visto en logs anteriores)
                var confirmBtns = document.querySelectorAll('input[type="submit"][value="Confirm"], button');
                for (var btn of confirmBtns) {
                    var text = (btn.value || btn.textContent || '').toLowerCase().trim();
                    if (text === 'confirm' || text === 'conferma') {
                        if (btn.offsetParent !== null && !btn.hasAttribute('data-dir')) {
                            btn.scrollIntoView({block: 'center'});
                            btn.click();
                            return 'Clicked CONFIRM button';
                        }
                    }
                }

                // PASO 1: Buscar input[type=submit] con value "Confirm" o similar
                var submits = document.querySelectorAll('input[type="submit"]');
                for (var sub of submits) {
                    var val = (sub.value || '').toLowerCase().trim();
                    // Priorizar Confirm, Add, Book
                    if (val === 'confirm' || val === 'conferma' ||
                        val.includes('add to cart') || val.includes('book') ||
                        val.includes('aggiungi') || val.includes('prenota')) {
                        if (sub.offsetParent !== null) {
                            sub.scrollIntoView({block: 'center'});
                            sub.click();
                            return 'Clicked submit: ' + sub.value;
                        }
                    }
                }

                // PASO 2: Buscar cualquier submit visible que NO sea SEARCH
                for (var sub of submits) {
                    var val = (sub.value || '').toLowerCase().trim();
                    if (val && val !== 'search' && val !== 'cerca' && sub.offsetParent !== null) {
                        sub.scrollIntoView({block: 'center'});
                        sub.click();
                        return 'Clicked other submit: ' + sub.value;
                    }
                }

                // PASO 3: Buscar boton con texto relevante
                var targetTexts = ['confirm', 'conferma', 'add to cart', 'add to basket', 'book now',
                                   'buy now', 'buy tickets', 'aggiungi al carrello', 'prenota ora',
                                   'acquista', 'procedi', 'checkout', 'proceed'];
                var excludeTexts = ['continue shopping', 'keep shopping', 'search', 'cerca', 'close'];

                var buttons = document.querySelectorAll('button:not([data-dir]), input[type="submit"], a.btn, a.button');

                for (var btn of buttons) {
                    if (btn.offsetParent === null) continue;
                    var text = (btn.textContent || btn.value || '').toLowerCase().trim();

                    var shouldExclude = false;
                    for (var ex of excludeTexts) {
                        if (text.includes(ex)) { shouldExclude = true; break; }
                    }
                    if (shouldExclude) continue;

                    for (var target of targetTexts) {
                        if (text.includes(target)) {
                            btn.scrollIntoView({block: 'center'});
                            btn.click();
                            return 'Clicked: ' + text.substring(0, 40);
                        }
                    }
                }

                // PASO 4: Buscar por clase especifica de ABC booking
                var classPatterns = ['.abc-book-btn', '.abc-add-to-cart', '.add-to-cart', '.btn-book',
                                     '.btn-primary', '.btn-success', '.book-button'];
                for (var pattern of classPatterns) {
                    try {
                        var btns = document.querySelectorAll(pattern);
                        for (var btn of btns) {
                            if (btn.offsetParent !== null && !btn.hasAttribute('data-dir')) {
                                var text = (btn.textContent || '').toLowerCase();
                                if (!text.includes('shopping') && !text.includes('search')) {
                                    btn.click();
                                    return 'Clicked pattern ' + pattern + ': ' + text.substring(0, 30);
                                }
                            }
                        }
                    } catch(e) {}
                }

                // Debug: listar todos los submits e inputs
                var debugInfo = [];
                var allSubmits = document.querySelectorAll('input[type="submit"]');
                for (var s of allSubmits) {
                    debugInfo.push('submit:' + (s.value || 'no-value') + '|visible:' + (s.offsetParent !== null));
                }
                return 'No confirm/add button. Submits: ' + debugInfo.join('; ');
            """)
            print(f"[Flow] Add to cart result: {add_to_cart_result}")
            time.sleep(5)  # Esperar a que procese

            # Verificar si hubo navegacion
            current_url = driver.current_url
            print(f"[Flow] URL despues de add-to-cart: {current_url}")

        except Exception as e:
            print(f"[Flow] Error en add-to-cart: {e}")

        # 6. Debug: ver todos los formularios
        print("[Flow] Analizando formularios en la pagina...")
        try:
            forms_info = driver.execute_script("""
                var forms = document.querySelectorAll('form');
                var info = [];
                forms.forEach(function(form, idx) {
                    var action = form.action || 'no-action';
                    var id = form.id || 'no-id';
                    var inputs = form.querySelectorAll('input, select, button').length;
                    info.push('Form' + idx + ': id=' + id + ', action=' + action.substring(0,50) + ', inputs=' + inputs);
                });
                return info.join(' | ');
            """)
            print(f"[Flow] Forms: {forms_info[:300] if forms_info else 'None'}")
        except Exception as e:
            print(f"[Flow] Error analizando forms: {e}")

        # 7. Explorar estructura de la pagina para encontrar como agregar al carrito
        print("[Flow] Explorando estructura de la pagina...")
        try:
            page_structure = driver.execute_script("""
                var result = {};

                // Buscar secciones de tarifa
                var tariffSections = document.querySelectorAll('.tariff, .tariff-row, .ticket-type, [class*="tariff"], [class*="ticket"]');
                result.tariffSections = tariffSections.length;

                // Buscar inputs de cantidad
                var qtyInputs = document.querySelectorAll('input[type="number"], input[name*="qty"], input[name*="quantity"]');
                result.qtyInputs = qtyInputs.length;

                // Buscar todos los elementos con data attributes relacionados
                var dataElements = document.querySelectorAll('[data-product], [data-item], [data-ticket], [data-tariff]');
                result.dataElements = dataElements.length;

                // Buscar la seccion activa despues de seleccionar slot
                var activeSection = document.querySelector('.slot-selected, .time-selected, .active-slot, [class*="selected"]');
                result.hasActiveSection = !!activeSection;

                // Buscar botones dentro de secciones de tarifa
                if (tariffSections.length > 0) {
                    var btnsInTariff = tariffSections[0].querySelectorAll('button');
                    result.btnsInFirstTariff = btnsInTariff.length;
                }

                // Buscar el HTML de la seccion principal
                var mainContent = document.querySelector('.event-content, .booking-section, main, #content');
                if (mainContent) {
                    result.mainContentPreview = mainContent.innerHTML.substring(0, 500);
                }

                return JSON.stringify(result, null, 2);
            """)
            print(f"[Flow] Page structure: {page_structure[:400] if page_structure else 'None'}")
        except Exception as e:
            print(f"[Flow] Error explorando pagina: {e}")

        # 7b. Analizar la estructura de tarifas en detalle
        print("[Flow] Analizando estructura de tarifas...")
        try:
            tariff_analysis = driver.execute_script("""
                // Buscar la primera tarifa con cantidad > 0
                var qtyInputs = document.querySelectorAll('input[type="number"]');
                for (var input of qtyInputs) {
                    var val = parseInt(input.value) || 0;
                    if (val > 0) {
                        // Encontramos una tarifa con cantidad
                        var row = input.closest('tr, .row, .tariff-row, div[class*="tariff"]');
                        if (row) {
                            return 'Found qty=' + val + ' in: ' + row.className + ' | HTML: ' + row.innerHTML.substring(0, 300);
                        }
                    }
                }

                // Si no hay cantidad > 0, mostrar el primer input
                if (qtyInputs.length > 0) {
                    var first = qtyInputs[0];
                    var row = first.closest('tr, .row, div');
                    return 'First qty input value=' + first.value + ' | Row HTML: ' + (row ? row.innerHTML.substring(0, 300) : 'no row');
                }

                return 'No qty inputs found';
            """)
            print(f"[Flow] Tariff analysis: {tariff_analysis[:400] if tariff_analysis else 'None'}")
        except Exception as e:
            print(f"[Flow] Error analizando tarifas: {e}")

        # 7c. Buscar boton de agregar especifico del sitio
        print("[Flow] Buscando boton agregar especifico...")
        try:
            add_cart_result = driver.execute_script("""
                // El sitio probablemente usa un boton con clase especifica o data attribute
                // Buscar todos los botones y mostrar sus atributos
                var allButtons = document.querySelectorAll('button');
                var buttonInfo = [];

                for (var i = 0; i < Math.min(allButtons.length, 50); i++) {
                    var btn = allButtons[i];
                    if (btn.offsetParent === null) continue; // Skip hidden

                    var info = {
                        idx: i,
                        text: (btn.textContent || '').trim().substring(0, 20),
                        class: (btn.className || '').substring(0, 30),
                        type: btn.type || '',
                        onclick: btn.getAttribute('onclick') ? 'yes' : 'no',
                        dataAttrs: []
                    };

                    // Capturar data attributes
                    for (var attr of btn.attributes) {
                        if (attr.name.startsWith('data-')) {
                            info.dataAttrs.push(attr.name + '=' + attr.value.substring(0, 20));
                        }
                    }

                    if (info.dataAttrs.length > 0 || info.onclick === 'yes' || info.text.length > 0) {
                        buttonInfo.push(info);
                    }
                }

                return JSON.stringify(buttonInfo.slice(0, 10));
            """)
            print(f"[Flow] Button analysis: {add_cart_result[:500] if add_cart_result else 'None'}")
        except Exception as e:
            print(f"[Flow] Error analizando botones: {e}")

        # 7d. Intentar click en boton con data-action o similar
        print("[Flow] Intentando click en boton de accion...")
        try:
            click_result = driver.execute_script("""
                // Buscar botones con data attributes de accion
                var actionButtons = document.querySelectorAll('button[data-action], button[data-add], button[data-submit], a[data-action]');
                if (actionButtons.length > 0) {
                    actionButtons[0].click();
                    return 'Clicked data-action button';
                }

                // Buscar por clase que contenga 'add' o 'submit' o 'buy'
                var classButtons = document.querySelectorAll('button[class*="add"], button[class*="submit"], button[class*="buy"], button[class*="cart"]');
                for (var btn of classButtons) {
                    if (btn.offsetParent !== null && !btn.hasAttribute('data-dir')) {
                        btn.click();
                        return 'Clicked class button: ' + btn.className;
                    }
                }

                // Buscar el boton principal de la pagina (generalmente el mas grande o con estilo primario)
                var primaryBtns = document.querySelectorAll('.btn-primary, .primary-button, button.main, button.submit');
                for (var btn of primaryBtns) {
                    if (btn.offsetParent !== null) {
                        btn.click();
                        return 'Clicked primary button';
                    }
                }

                return 'No action button found';
            """)
            print(f"[Flow] Click result: {click_result}")
            time.sleep(10)
        except Exception as e:
            print(f"[Flow] Error clicking: {e}")

        # 8. Manejar posibles alerts
        print("[Flow] Verificando alerts...")
        try:
            from selenium.webdriver.common.alert import Alert
            alert = Alert(driver)
            alert_text = alert.text
            print(f"[Flow] Alert encontrado: {alert_text}")
            alert.dismiss()  # Cancelar para no vaciar el carrito
            print("[Flow] Alert cancelado")
            time.sleep(2)
        except:
            print("[Flow] No hay alerts pendientes")

        # 9. Verificar si navegamos al carrito y obtener mas cookies
        print("[Flow] Verificando navegacion al carrito...")
        try:
            current_url = driver.current_url
            print(f"[Flow] URL actual: {current_url}")

            # Si estamos en el carrito, bien! Si no, NO navegar para no perder sesion
            if 'cart' in current_url or 'carrello' in current_url:
                print("[Flow] Ya estamos en el carrito!")
            else:
                print("[Flow] No estamos en carrito, verificando si hay items...")
                # Verificar si hay indicador de carrito con items
                cart_count = driver.execute_script("""
                    var cartBadge = document.querySelector('.cart-count, .badge, .cart-items-count');
                    return cartBadge ? cartBadge.textContent : '0';
                """)
                print(f"[Flow] Items en carrito: {cart_count}")
        except Exception as e:
            print(f"[Flow] Error verificando carrito: {e}")

        # 5c. Segundo intento - si no navegamos al carrito, intentar de nuevo
        if 'cart' not in current_url and 'carrello' not in current_url:
            print("[Flow] SEGUNDO INTENTO: Buscando boton Confirm nuevamente...")
            time.sleep(3)
            try:
                retry_result = driver.execute_script("""
                    // Buscar especificamente input[type=submit][value=Confirm]
                    var confirms = document.querySelectorAll('input[type="submit"]');
                    for (var c of confirms) {
                        var val = (c.value || '').trim();
                        console.log('Found submit: ' + val + ', visible: ' + (c.offsetParent !== null));
                        if (val.toLowerCase() === 'confirm' && c.offsetParent !== null) {
                            c.scrollIntoView({block: 'center'});
                            c.click();
                            return 'RETRY: Clicked Confirm';
                        }
                    }

                    // Intentar navegar directamente al checkout
                    var checkoutLinks = document.querySelectorAll('a[href*="checkout"], a[href*="cart"]');
                    for (var link of checkoutLinks) {
                        var text = (link.textContent || '').toLowerCase();
                        if (!text.includes('empty') && !text.includes('shopping') && link.offsetParent !== null) {
                            link.click();
                            return 'RETRY: Clicked checkout/cart link';
                        }
                    }

                    return 'RETRY: No confirm button found';
                """)
                print(f"[Flow] Retry result: {retry_result}")
                time.sleep(5)

                # Verificar URL despues del retry
                current_url = driver.current_url
                print(f"[Flow] URL despues de retry: {current_url}")

            except Exception as e:
                print(f"[Flow] Error en retry: {e}")

        # 5d. Ir directamente al carrito para forzar sesion
        print("[Flow] Navegando directamente al carrito...")
        try:
            driver.get("https://ticketing.colosseo.it/en/cart/")
            time.sleep(5)
            cart_url = driver.current_url
            print(f"[Flow] URL carrito final: {cart_url}")

            # Verificar si hay items
            cart_items = driver.execute_script("""
                var body = document.body.innerText;
                if (body.includes('empty') || body.includes('vuoto') || body.includes('vacío')) {
                    return 'EMPTY';
                }
                if (body.includes('item') || body.includes('ticket') || body.includes('bigliett')) {
                    return 'HAS_ITEMS';
                }
                return 'UNKNOWN: ' + body.substring(0, 100);
            """)
            print(f"[Flow] Estado carrito: {cart_items}")

        except Exception as e:
            print(f"[Flow] Error navegando al carrito: {e}")

        # 6. Forzar multiples peticiones AJAX para generar cookies de sesion
        print("[Flow] Forzando peticiones AJAX...")
        try:
            # Llamada al calendario
            ajax_result = driver.execute_script("""
                return new Promise((resolve) => {
                    var formData = new FormData();
                    formData.append('action', 'mtajax_calendars_month');
                    formData.append('guids[entranceEvent_guid][]', 'a9a4b0f8-bf3c-4f22-afcd-196a27be04b9');
                    formData.append('singleDaySession', 'false');
                    formData.append('month', new Date().getMonth() + 1);
                    formData.append('year', new Date().getFullYear());

                    fetch('/mtajax/calendars_month', {
                        method: 'POST',
                        body: formData,
                        credentials: 'include'
                    })
                    .then(r => r.text())
                    .then(t => resolve('Calendar: ' + t.substring(0, 50)))
                    .catch(e => resolve('Error: ' + e.message));
                });
            """)
            print(f"[Flow] AJAX calendar: {ajax_result[:80] if ajax_result else 'None'}")
            time.sleep(2)

            # Llamada al carrito
            cart_result = driver.execute_script("""
                return new Promise((resolve) => {
                    fetch('/en/cart/', {
                        method: 'GET',
                        credentials: 'include'
                    })
                    .then(r => r.text())
                    .then(t => resolve('Cart loaded: ' + t.length + ' chars'))
                    .catch(e => resolve('Error: ' + e.message));
                });
            """)
            print(f"[Flow] AJAX cart: {cart_result[:80] if cart_result else 'None'}")
            time.sleep(2)

            # Llamada a wp-admin/admin-ajax.php (endpoint comun de WordPress)
            wp_result = driver.execute_script("""
                return new Promise((resolve) => {
                    var formData = new FormData();
                    formData.append('action', 'abc_get_cart');

                    fetch('/wp-admin/admin-ajax.php', {
                        method: 'POST',
                        body: formData,
                        credentials: 'include'
                    })
                    .then(r => r.text())
                    .then(t => resolve('WP: ' + t.substring(0, 50)))
                    .catch(e => resolve('Error: ' + e.message));
                });
            """)
            print(f"[Flow] AJAX wp: {wp_result[:80] if wp_result else 'None'}")
            time.sleep(2)

        except Exception as e:
            print(f"[Flow] Error AJAX: {e}")

        return True

    except Exception as e:
        print(f"[Flow] Error general: {e}")
        return False


def get_cookies_via_cdp(driver):
    """Obtiene TODAS las cookies usando Chrome DevTools Protocol"""
    print("[CDP] Obteniendo cookies via DevTools Protocol...")
    cookies_found = []

    try:
        # Metodo 1: Network.getAllCookies (obtiene todas las cookies del navegador)
        try:
            result = driver.execute_cdp_cmd('Network.getAllCookies', {})
            cdp_cookies = result.get('cookies', [])
            print(f"[CDP] Network.getAllCookies: {len(cdp_cookies)} cookies")

            for c in cdp_cookies:
                cookies_found.append({
                    'name': c.get('name'),
                    'value': c.get('value'),
                    'domain': c.get('domain', '.colosseo.it'),
                    'path': c.get('path', '/'),
                    'secure': c.get('secure', False),
                    'httpOnly': c.get('httpOnly', False)
                })
                print(f"  - {c.get('name')} (httpOnly: {c.get('httpOnly', False)})")
        except Exception as e:
            print(f"[CDP] Error Network.getAllCookies: {e}")

        # Metodo 2: Storage.getCookies para dominio especifico
        try:
            result = driver.execute_cdp_cmd('Storage.getCookies', {
                'browserContextId': None
            })
            storage_cookies = result.get('cookies', [])
            print(f"[CDP] Storage.getCookies: {len(storage_cookies)} cookies")
        except Exception as e:
            print(f"[CDP] Storage.getCookies no disponible: {e}")

        # Metodo 3: Page.getCookies (cookies de la pagina actual)
        try:
            result = driver.execute_cdp_cmd('Page.getCookies', {})
            page_cookies = result.get('cookies', [])
            print(f"[CDP] Page.getCookies: {len(page_cookies)} cookies")

            # Agregar las que no tengamos
            existing_names = {c['name'] for c in cookies_found}
            for c in page_cookies:
                if c.get('name') not in existing_names:
                    cookies_found.append({
                        'name': c.get('name'),
                        'value': c.get('value'),
                        'domain': c.get('domain', '.colosseo.it'),
                        'path': c.get('path', '/'),
                        'secure': c.get('secure', False),
                        'httpOnly': c.get('httpOnly', False)
                    })
                    print(f"  + {c.get('name')} (from Page.getCookies)")
        except Exception as e:
            print(f"[CDP] Page.getCookies no disponible: {e}")

    except Exception as e:
        print(f"[CDP] Error general: {e}")

    return cookies_found


def get_cookies_from_network_logs(driver):
    """Extrae cookies HttpOnly de los logs de red de Chrome"""
    import json as json_module

    print("[Network] Analizando logs de red para cookies HttpOnly...")
    cookies_found = {}
    set_cookie_count = 0

    try:
        logs = driver.get_log('performance')
        print(f"[Network] Procesando {len(logs)} entradas de log...")

        for entry in logs:
            try:
                log = json_module.loads(entry['message'])['message']
                method = log.get('method', '')

                # Buscar respuestas con Set-Cookie
                if method == 'Network.responseReceivedExtraInfo':
                    headers = log.get('params', {}).get('headers', {})
                    for key, value in headers.items():
                        if key.lower() == 'set-cookie':
                            set_cookie_count += 1
                            # Parsear la cookie
                            cookie_parts = value.split(';')[0].split('=', 1)
                            if len(cookie_parts) == 2:
                                name, val = cookie_parts
                                cookies_found[name.strip()] = val.strip()
                                # Debug: mostrar cookies importantes
                                if 'php' in name.lower() or 'octofence' in name.lower() or 'waap' in name.lower():
                                    print(f"[Network] IMPORTANTE encontrada: {name}")

                # Buscar cookies en requests salientes
                if method == 'Network.requestWillBeSentExtraInfo':
                    headers = log.get('params', {}).get('headers', {})
                    cookie_header = headers.get('Cookie', headers.get('cookie', ''))
                    if cookie_header:
                        for cookie_pair in cookie_header.split(';'):
                            if '=' in cookie_pair:
                                name, val = cookie_pair.split('=', 1)
                                name = name.strip()
                                if name not in cookies_found:
                                    cookies_found[name] = val.strip()

                # Buscar tambien en Network.responseReceived para headers
                if method == 'Network.responseReceived':
                    response = log.get('params', {}).get('response', {})
                    headers = response.get('headers', {})
                    for key, value in headers.items():
                        if key.lower() == 'set-cookie':
                            cookie_parts = value.split(';')[0].split('=', 1)
                            if len(cookie_parts) == 2:
                                name, val = cookie_parts
                                cookies_found[name.strip()] = val.strip()

            except Exception as e:
                continue

        print(f"[Network] Set-Cookie headers encontrados: {set_cookie_count}")
        print(f"[Network] Cookies unicas extraidas: {len(cookies_found)}")
        for name in sorted(cookies_found.keys()):
            is_critical = 'php' in name.lower() or 'waap' in name.lower()
            print(f"  {'*' if is_critical else '-'} {name}")

    except Exception as e:
        print(f"[Network] Error procesando logs: {e}")

    return cookies_found


def get_cookies(driver):
    """Obtiene todas las cookies del navegador"""
    # Verificar contenido de la pagina actual
    try:
        page_source = driver.page_source
        print(f"[Debug] Page length: {len(page_source)}")
        if 'empty' in page_source.lower() or 'vacío' in page_source.lower() or 'vuoto' in page_source.lower():
            print("[Debug] El carrito parece estar vacio")
        if 'cart' in page_source.lower() or 'carrello' in page_source.lower():
            print("[Debug] Pagina contiene referencia a carrito")

        # Verificar si hay items en el carrito
        if 'There\'s an item in your cart' in page_source or 'items in your cart' in page_source:
            print("[Debug] HAY ITEMS EN EL CARRITO!")
    except:
        pass

    # METODO 1: CDP - obtiene TODAS las cookies incluyendo HttpOnly
    cdp_cookies = get_cookies_via_cdp(driver)

    # METODO 2: Network logs (backup)
    network_cookies = get_cookies_from_network_logs(driver)

    # METODO 3: Selenium standard (backup)
    cookies = driver.get_cookies()
    print(f"[Cookies] Obtenidas {len(cookies)} cookies via driver.get_cookies()")

    # Debug: mostrar cookies del driver
    print("[Debug] Cookies de driver.get_cookies():")
    for c in cookies:
        print(f"  - {c['name']} (domain: {c.get('domain', 'N/A')})")

    # PRIORIDAD 1: Usar cookies de CDP (incluye HttpOnly)
    relevant = []
    important_names = ['PHPSESSID', 'octofence', 'waap']

    # Filtrar cookies CDP relevantes
    for c in cdp_cookies:
        domain = c.get('domain', '')
        name = c.get('name', '')

        is_relevant = (
            'colosseo' in domain or
            'ticketing' in domain or
            any(imp in name.lower() for imp in important_names)
        )

        if is_relevant:
            relevant.append(c)

    print(f"[Cookies] Relevantes de CDP: {len(relevant)}")

    # PRIORIDAD 2: Agregar cookies del driver que no esten en CDP
    existing_names = {c['name'] for c in relevant}
    for c in cookies:
        domain = c.get('domain', '')
        name = c.get('name', '')

        is_relevant = (
            'colosseo' in domain or
            'ticketing' in domain or
            any(imp in name.lower() for imp in important_names)
        )

        if is_relevant and name not in existing_names:
            relevant.append({
                'name': c['name'],
                'value': c['value'],
                'domain': c.get('domain', '.colosseo.it'),
                'path': c.get('path', '/'),
                'secure': c.get('secure', False),
                'httpOnly': c.get('httpOnly', False)
            })
            print(f"  + {name} (from driver)")

    # PRIORIDAD 3: Agregar cookies de network logs
    existing_names = {c['name'] for c in relevant}
    for name, value in network_cookies.items():
        if name not in existing_names:
            relevant.append({
                'name': name,
                'value': value,
                'domain': '.colosseo.it',
                'path': '/',
                'secure': True,
                'httpOnly': True
            })
            print(f"  + {name} (from network logs)")

    print(f"[Cookies] Total combinadas: {len(relevant)}")

    # Mostrar resumen de cookies criticas
    cookie_names = [c['name'] for c in relevant]
    critical = ['PHPSESSID', 'octofence-waap-id', 'octofence-waap-sessid']
    print("[Cookies] Cookies criticas:")
    for crit in critical:
        found = crit in cookie_names
        print(f"  - {crit}: {'SI' if found else 'NO'}")

    return relevant


def save_to_supabase(cookies):
    """Guarda cookies en Supabase Storage"""
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("[Supabase] No configurado")
        return False

    try:
        from supabase import create_client

        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

        # Incluir info del proxy usado para que Vercel use el mismo
        proxy_info = None
        if PROXY_HOST and PROXY_PORT:
            proxy_info = {
                "host": PROXY_HOST,
                "port": PROXY_PORT,
                "user": PROXY_USER if PROXY_USER else None,
                "url": f"http://{PROXY_USER}:{PROXY_PASS}@{PROXY_HOST}:{PROXY_PORT}" if PROXY_USER else f"http://{PROXY_HOST}:{PROXY_PORT}"
            }

        data = {
            "cookies": cookies,
            "timestamp": datetime.utcnow().isoformat() + 'Z',
            "source": "railway-uc",
            "proxy": proxy_info
        }

        cookies_json = json.dumps(data, indent=2).encode('utf-8')

        bucket = 'colosseo-files'
        path = 'cookies/cookies_auto.json'

        # Crear bucket si no existe
        try:
            supabase.storage.get_bucket(bucket)
        except:
            try:
                supabase.storage.create_bucket(bucket, options={'public': False})
            except:
                pass

        # Eliminar archivo anterior
        try:
            supabase.storage.from_(bucket).remove([path])
        except:
            pass

        # Subir
        supabase.storage.from_(bucket).upload(
            path, cookies_json,
            file_options={"content-type": "application/json", "upsert": "true"}
        )

        print("[Supabase] Cookies guardadas exitosamente")
        return True

    except Exception as e:
        print(f"[Supabase] Error: {e}")
        return False


def test_api_with_cookies(driver, cookies):
    """Prueba hacer una consulta API usando el navegador"""
    print("\n[API Test] Probando consulta con cookies...")

    try:
        # Hacer la consulta AJAX desde el navegador
        result = driver.execute_script("""
            return new Promise((resolve) => {
                var formData = new FormData();
                formData.append('action', 'mtajax_calendars_month');
                formData.append('guids[entranceEvent_guid][]', 'a9a4b0f8-bf3c-4f22-afcd-196a27be04b9');
                formData.append('singleDaySession', 'false');
                formData.append('month', new Date().getMonth() + 1);
                formData.append('year', new Date().getFullYear());

                fetch('/mtajax/calendars_month', {
                    method: 'POST',
                    body: formData,
                    credentials: 'include'
                })
                .then(r => r.json())
                .then(data => {
                    if (data && data.timeslots) {
                        resolve({success: true, timeslots: data.timeslots.length, sample: data.timeslots.slice(0, 2)});
                    } else if (data && data.message) {
                        resolve({success: false, error: data.message});
                    } else {
                        resolve({success: false, error: 'Unknown response', keys: Object.keys(data || {})});
                    }
                })
                .catch(e => resolve({success: false, error: e.message}));
            });
        """)

        print(f"[API Test] Result: {result}")

        if result and result.get('success'):
            print(f"[API Test] SUCCESS! {result.get('timeslots')} timeslots found")
            return True
        else:
            print(f"[API Test] Failed: {result.get('error', 'unknown')}")
            return False

    except Exception as e:
        print(f"[API Test] Error: {e}")
        return False


def fetch_availability_from_browser(driver):
    """
    Consulta la disponibilidad usando las cookies del navegador.

    ESTRATEGIA:
    1. Navegar a la página del tour - esto carga el calendario automáticamente
    2. Capturar la respuesta de calendars_month de los network logs
    3. Si no funciona, intentar AJAX directo
    """
    import json as json_module

    print("\n[Availability] Consultando disponibilidad via API...")

    TOURS = {
        "24h-grupos": {
            "nombre": "24h Colosseo, Foro Romano y Palatino - GRUPOS",
            "guid": "a9a4b0f8-bf3c-4f22-afcd-196a27be04b9",
            "url": "https://ticketing.colosseo.it/en/eventi/24h-colosseo-foro-romano-palatino-gruppi/"
        },
        "arena": {
            "nombre": "Colosseo con ACCESO A LA ARENA",
            "guid": "8d1c991c-a15f-42bc-8cb5-bd738aa19c70",
            "url": "https://ticketing.colosseo.it/en/eventi/colosseo-arena/"
        }
    }

    all_results = {}

    for tour_key, tour_info in TOURS.items():
        print(f"\n[Availability] Tour: {tour_info['nombre'][:50]}...")

        tour_data = {
            "nombre": tour_info['nombre'],
            "guid": tour_info['guid'],
            "timeslots": [],
            "fechas_disponibles": 0,
            "total_plazas": 0
        }

        try:
            # PASO 1: Navegar a la página del tour para que cargue el calendario
            print(f"  Navegando a pagina del tour...")
            driver.get(tour_info['url'])
            time.sleep(8)

            # PASO 2: Capturar respuesta del calendario desde network logs
            print("  Capturando respuesta del calendario...")
            logs = driver.get_log('performance')
            print(f"  Network logs: {len(logs)} entradas")

            for entry in logs:
                try:
                    log = json_module.loads(entry['message'])['message']
                    method = log.get('method', '')

                    if method == 'Network.responseReceived':
                        response = log.get('params', {}).get('response', {})
                        url = response.get('url', '')

                        if 'calendars_month' in url:
                            request_id = log.get('params', {}).get('requestId')
                            try:
                                body = driver.execute_cdp_cmd('Network.getResponseBody', {'requestId': request_id})
                                body_text = body.get('body', '')
                                if body_text:
                                    response_data = json_module.loads(body_text)

                                    # Buscar timeslots
                                    timeslots = None
                                    if 'timeslots' in response_data:
                                        timeslots = response_data.get('timeslots', [])
                                    elif 'data' in response_data:
                                        inner = response_data.get('data', {})
                                        if isinstance(inner, dict) and 'timeslots' in inner:
                                            timeslots = inner.get('timeslots', [])
                                        elif isinstance(inner, list):
                                            timeslots = inner

                                    if timeslots:
                                        tour_data['timeslots'].extend(timeslots)
                                        print(f"    +{len(timeslots)} timeslots desde network logs")
                            except:
                                pass
                except:
                    continue

            # PASO 3: Si capturamos timeslots, intentar AJAX para meses adicionales
            if tour_data['timeslots']:
                print(f"  Capturados {len(tour_data['timeslots'])} timeslots del mes actual")
                print("  Intentando obtener meses adicionales via AJAX...")

            # Consultar meses adicionales
            months_to_fetch = 6
            current_date = datetime.now()

            for month_offset in range(months_to_fetch):
                target_month = current_date.month + month_offset
                target_year = current_date.year

                while target_month > 12:
                    target_month -= 12
                    target_year += 1

                print(f"  Consultando {target_year}-{target_month:02d}...")

                # Usar XMLHttpRequest con los headers exactos que usa el sitio
                ajax_result = driver.execute_script(f"""
                    return new Promise((resolve) => {{
                        try {{
                            var xhr = new XMLHttpRequest();
                            xhr.open('POST', 'https://ticketing.colosseo.it/mtajax/calendars_month', true);
                            xhr.setRequestHeader('Content-Type', 'application/x-www-form-urlencoded; charset=UTF-8');
                            xhr.setRequestHeader('X-Requested-With', 'XMLHttpRequest');
                            xhr.setRequestHeader('Accept', 'application/json, text/javascript, */*; q=0.01');
                            xhr.withCredentials = true;

                            xhr.onload = function() {{
                                try {{
                                    var data = JSON.parse(xhr.responseText);

                                    // Buscar timeslots en diferentes ubicaciones de la respuesta
                                    var timeslots = null;
                                    if (data && data.timeslots) {{
                                        timeslots = data.timeslots;
                                    }} else if (data && data.data) {{
                                        if (Array.isArray(data.data)) {{
                                            timeslots = data.data;
                                        }} else if (data.data && data.data.timeslots) {{
                                            timeslots = data.data.timeslots;
                                        }}
                                    }}

                                    if (timeslots && timeslots.length > 0) {{
                                        resolve({{
                                            success: true,
                                            timeslots: timeslots,
                                            count: timeslots.length
                                        }});
                                    }} else if (data && data.message) {{
                                        resolve({{success: false, error: data.message, raw: JSON.stringify(data).substring(0, 200)}});
                                    }} else {{
                                        resolve({{success: false, error: 'No timeslots', keys: Object.keys(data || {{}}), raw: JSON.stringify(data).substring(0, 200)}});
                                    }}
                                }} catch(e) {{
                                    resolve({{success: false, error: 'Parse: ' + e.message, raw: xhr.responseText.substring(0, 200)}});
                                }}
                            }};

                            xhr.onerror = function() {{
                                resolve({{success: false, error: 'Network error'}});
                            }};

                            var params = 'action=mtajax_calendars_month' +
                                         '&guids%5BentranceEvent_guid%5D%5B%5D={tour_info["guid"]}' +
                                         '&singleDaySession=false' +
                                         '&month={target_month}' +
                                         '&year={target_year}';

                            xhr.send(params);
                        }} catch(e) {{
                            resolve({{success: false, error: 'Exception: ' + e.message}});
                        }}
                    }});
                """)

                if ajax_result and ajax_result.get('success'):
                    timeslots = ajax_result.get('timeslots', [])
                    tour_data['timeslots'].extend(timeslots)
                    print(f"    +{ajax_result.get('count', 0)} timeslots")
                else:
                    error = ajax_result.get('error', 'unknown') if ajax_result else 'null'
                    raw = ajax_result.get('raw', '') if ajax_result else ''
                    print(f"    Error: {str(error)[:50]}")
                    if raw:
                        print(f"    Raw: {raw[:100]}")

                time.sleep(0.5)  # Pequeña pausa entre requests

        except Exception as e:
            print(f"  Error: {str(e)[:50]}")

        # Procesar timeslots para calcular totales
        fechas_con_disponibilidad = set()
        total_plazas = 0
        for ts in tour_data['timeslots']:
            capacity = ts.get('capacity', 0)
            if capacity > 0:
                fecha = ts.get('startDateTime', '')[:10]
                fechas_con_disponibilidad.add(fecha)
                total_plazas += capacity

        tour_data['fechas_disponibles'] = len(fechas_con_disponibilidad)
        tour_data['total_plazas'] = total_plazas

        all_results[tour_key] = tour_data
        print(f"  Total: {tour_data['fechas_disponibles']} fechas, {tour_data['total_plazas']} plazas")

    return all_results


def save_availability_to_supabase(availability_data):
    """Guarda los datos de disponibilidad en Supabase Storage"""
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("[Supabase] No configurado para disponibilidad")
        return False

    try:
        from supabase import create_client

        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

        data = {
            "availability": availability_data,
            "timestamp": datetime.utcnow().isoformat() + 'Z',
            "source": "railway-browser"
        }

        availability_json = json.dumps(data, indent=2).encode('utf-8')

        bucket = 'colosseo-files'
        path = 'availability/availability_cache.json'

        # Eliminar archivo anterior
        try:
            supabase.storage.from_(bucket).remove([path])
        except:
            pass

        # Subir
        supabase.storage.from_(bucket).upload(
            path, availability_json,
            file_options={"content-type": "application/json", "upsert": "true"}
        )

        print("[Supabase] Disponibilidad guardada exitosamente")

        # También actualizar el histórico Excel
        update_historico_excel(supabase, availability_data)

        return True

    except Exception as e:
        print(f"[Supabase] Error guardando disponibilidad: {e}")
        return False


def update_historico_excel(supabase, availability_data):
    """
    Actualiza el archivo histórico Excel en Supabase.
    Descarga el existente, agrega nueva columna con datos actuales, y lo sube.
    """
    from io import BytesIO

    try:
        import pandas as pd
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        print(f"[Historico] Dependencias no disponibles: {e}")
        return False

    bucket = 'colosseo-files'
    path = 'historico/historico_disponibilidad.xlsx'
    # Convertir timestamp a hora de Roma
    from zoneinfo import ZoneInfo
    rome_now = datetime.now(ZoneInfo('Europe/Rome'))
    timestamp = rome_now.strftime("%Y-%m-%d %H:%M")

    print(f"\n[Historico] Actualizando historico Excel...")

    try:
        # Intentar descargar archivo existente
        try:
            existing_data = supabase.storage.from_(bucket).download(path)
            wb = load_workbook(BytesIO(existing_data))
            print(f"[Historico] Archivo existente descargado ({len(existing_data)} bytes)")
        except Exception as e:
            print(f"[Historico] No hay archivo existente, creando nuevo: {e}")
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

        # Procesar cada tour
        for tour_key, tour_data in availability_data.items():
            sheet_name = tour_key[:31]  # Max 31 chars para nombre de hoja
            timeslots = tour_data.get('timeslots', [])

            if not timeslots:
                print(f"[Historico] {tour_key}: sin timeslots")
                continue

            # Crear diccionario de (fecha, hora) -> capacidad
            datos_actuales = {}
            for ts in timeslots:
                start = ts.get('startDateTime', '')
                if not start:
                    continue

                # Parsear fecha y hora (formato: 2025-12-04T07:30:00Z)
                # Convertir de UTC a hora de Roma (+1h en invierno, +2h en verano)
                try:
                    from zoneinfo import ZoneInfo
                    utc_dt = datetime.fromisoformat(start.replace('Z', '+00:00'))
                    rome_dt = utc_dt.astimezone(ZoneInfo('Europe/Rome'))
                    fecha = rome_dt.strftime('%Y-%m-%d')
                    hora = rome_dt.strftime('%H:%M')
                    capacidad = ts.get('capacity', 0)
                    capacidad_original = ts.get('originalCapacity', capacidad)

                    key = (fecha, hora)
                    datos_actuales[key] = {
                        'capacidad': capacidad,
                        'capacidad_original': capacidad_original
                    }
                except:
                    continue

            if not datos_actuales:
                continue

            # Crear o actualizar hoja
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)

                # Headers fijos
                ws['A1'] = 'Fecha'
                ws['B1'] = 'Hora'
                ws['C1'] = 'Capacidad Total'

                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)

                for col in ['A1', 'B1', 'C1']:
                    ws[col].fill = header_fill
                    ws[col].font = header_font
                    ws[col].alignment = Alignment(horizontal='center')

                ws.column_dimensions['A'].width = 12
                ws.column_dimensions['B'].width = 8
                ws.column_dimensions['C'].width = 14

                # Escribir filas iniciales
                row = 2
                for (fecha, hora), datos in sorted(datos_actuales.items()):
                    ws[f'A{row}'] = fecha
                    ws[f'B{row}'] = hora
                    ws[f'C{row}'] = datos['capacidad_original']
                    row += 1

                col_num = 4
            else:
                ws = wb[sheet_name]

                # Leer filas existentes
                filas_existentes = {}
                for row in range(2, ws.max_row + 1):
                    fecha_cell = ws[f'A{row}'].value
                    hora_cell = ws[f'B{row}'].value
                    if fecha_cell and hora_cell:
                        key = (str(fecha_cell), str(hora_cell))
                        filas_existentes[key] = row

                # Agregar horarios nuevos
                next_row = ws.max_row + 1
                for (fecha, hora) in sorted(datos_actuales.keys()):
                    if (fecha, hora) not in filas_existentes:
                        ws[f'A{next_row}'] = fecha
                        ws[f'B{next_row}'] = hora
                        ws[f'C{next_row}'] = datos_actuales[(fecha, hora)]['capacidad_original']
                        filas_existentes[(fecha, hora)] = next_row
                        next_row += 1

                col_num = ws.max_column + 1

            # Agregar columna con timestamp actual
            col_letter = get_column_letter(col_num)

            # Header
            ws[f'{col_letter}1'] = timestamp
            ws[f'{col_letter}1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            ws[f'{col_letter}1'].font = Font(color='FFFFFF', bold=True)
            ws[f'{col_letter}1'].alignment = Alignment(horizontal='center')
            ws.column_dimensions[col_letter].width = 14

            # Llenar datos
            for row in range(2, ws.max_row + 1):
                fecha_cell = ws[f'A{row}'].value
                hora_cell = ws[f'B{row}'].value

                if fecha_cell and hora_cell:
                    key = (str(fecha_cell), str(hora_cell))

                    if key in datos_actuales:
                        capacidad = datos_actuales[key]['capacidad']
                        capacidad_orig = datos_actuales[key]['capacidad_original']
                    else:
                        capacidad = '-'
                        capacidad_orig = 0

                    cell = ws[f'{col_letter}{row}']
                    cell.value = capacidad
                    cell.alignment = Alignment(horizontal='center')

                    # Formato condicional
                    if capacidad == '-':
                        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                        cell.font = Font(color='999999', italic=True)
                    elif capacidad == 0:
                        cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                        cell.font = Font(color='FFFFFF', bold=True)
                    elif capacidad_orig > 0:
                        porcentaje = (capacidad / capacidad_orig) * 100
                        if porcentaje < 30:
                            cell.fill = PatternFill(start_color='FFE066', end_color='FFE066', fill_type='solid')
                        elif porcentaje > 70:
                            cell.fill = PatternFill(start_color='95E1D3', end_color='95E1D3', fill_type='solid')

            print(f"[Historico] {sheet_name}: {len(datos_actuales)} horarios actualizados")

        # Guardar y subir
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_bytes = output.getvalue()

        # Eliminar archivo anterior
        try:
            supabase.storage.from_(bucket).remove([path])
        except:
            pass

        # Subir
        supabase.storage.from_(bucket).upload(
            path, file_bytes,
            file_options={"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "upsert": "true"}
        )

        print(f"[Historico] Archivo subido exitosamente ({len(file_bytes)} bytes)")
        return True

    except Exception as e:
        print(f"[Historico] Error actualizando historico: {e}")
        import traceback
        traceback.print_exc()
        return False


def try_with_proxy(proxy=None, attempt=1):
    """
    Intenta obtener cookies con un proxy específico.
    Retorna (success, cookies, driver) o (False, None, None) si falla.
    """
    driver = None
    try:
        driver = setup_driver(proxy_override=proxy)

        # Navegar a la pagina
        print(f"\n[Navigate] Abriendo {TOUR_URL[:50]}...")
        driver.get(TOUR_URL)

        # Esperar Octofence - si falla, probablemente IP bloqueada
        if not wait_for_octofence(driver, timeout=60):
            print(f"[Attempt {attempt}] Octofence no pasado - posible bloqueo de IP")
            return False, None, driver

        # Completar flujo de reserva
        time.sleep(3)
        complete_booking_flow(driver)

        # Esperar a que se generen cookies de sesion
        print("\n[Wait] Esperando generacion de cookies...")
        time.sleep(5)

        # Obtener cookies
        cookies = get_cookies(driver)

        if len(cookies) >= 5:
            return True, cookies, driver
        else:
            print(f"[Attempt {attempt}] Solo {len(cookies)} cookies obtenidas")
            return False, None, driver

    except Exception as e:
        print(f"[Attempt {attempt}] Error: {e}")
        return False, None, driver


def main():
    print("=" * 60)
    print("COOKIE FETCHER - Railway + undetected-chromedriver")
    print(f"Timestamp: {datetime.now().isoformat()}")
    print("=" * 60)

    # Inicializar proxies desde Webshare API
    initialize_proxies()

    # Mostrar proxies disponibles
    print(f"\n[Proxy] {len(PROXIES)} proxies configurados")
    for i, p in enumerate(PROXIES):
        country = p.get('country', 'XX')
        print(f"  {i+1}. {p['host']}:{p['port']} ({country})")
    sys.stdout.flush()

    # Iniciar display virtual
    start_xvfb()

    max_attempts = max(len(PROXIES), 1) if PROXIES else 1
    driver = None
    cookies = None
    success = False

    for attempt in range(1, max_attempts + 1):
        print(f"\n{'='*40}")
        print(f"INTENTO {attempt}/{max_attempts}")
        print(f"{'='*40}")
        sys.stdout.flush()

        # Cerrar driver anterior si existe
        if driver:
            try:
                driver.quit()
            except:
                pass
            driver = None

        # Intentar con el proxy actual
        success, cookies, driver = try_with_proxy(attempt=attempt)

        if success:
            print(f"\n[Success] Intento {attempt} exitoso!")
            break
        else:
            # Rotar al siguiente proxy
            if len(PROXIES) > 1:
                next_proxy = rotate_proxy()
                print(f"[Retry] Rotando al siguiente proxy...")
            else:
                print(f"[Retry] No hay más proxies para probar")
                break

    if not success:
        print("\n" + "=" * 60)
        print("RESULTADO: FALLO - Todos los proxies bloqueados")
        print("=" * 60)
        if driver:
            try:
                driver.quit()
            except:
                pass
        return 1

    # Proceso exitoso - continuar con guardar cookies y disponibilidad
    try:
        print(f"\n[Success] {len(cookies)} cookies obtenidas")
        sys.stdout.flush()

        # Verificar cookies criticas
        names = [c['name'] for c in cookies]
        critical = ['PHPSESSID', 'octofence-waap-id', 'octofence-waap-sessid']
        has_critical = any(c in names for c in critical)

        if has_critical:
            print("[Success] Cookies criticas encontradas!")
        else:
            print("[Warning] Faltan algunas cookies criticas, pero intentando guardar...")
        sys.stdout.flush()

        # Guardar cookies en Supabase PRIMERO (antes de disponibilidad)
        print("\n[Supabase] Guardando cookies...")
        sys.stdout.flush()
        save_to_supabase(cookies)

        # Consultar disponibilidad completa desde el navegador
        print("\n[Availability] Iniciando consulta de disponibilidad...")
        sys.stdout.flush()
        try:
            availability = fetch_availability_from_browser(driver)
            if availability:
                save_availability_to_supabase(availability)
            else:
                print("[Availability] No se pudo obtener disponibilidad")
        except Exception as e:
            print(f"[Availability] ERROR: {e}")
            import traceback
            traceback.print_exc()
        sys.stdout.flush()

        # Guardar local como backup
        with open('cookies_colosseo.json', 'w') as f:
            json.dump(cookies, f, indent=2)
        print("[Local] Backup guardado")

        print("\n" + "=" * 60)
        print("RESULTADO: EXITO")
        print("=" * 60)
        return 0

    except Exception as e:
        print(f"\n[Error] {e}")
        import traceback
        traceback.print_exc()
        return 1

    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass
            print("[Driver] Cerrado")


if __name__ == "__main__":
    sys.exit(main())
